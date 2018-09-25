#!/usr/bin/env python

from sys import argv
from pprint import pprint
from openpyxl import Workbook, load_workbook

from libs.transformations import COLUMN_TRANSF_RULES
from libs.color import get_color
from libs.utilities import *

# Maurice Laxroix Strap|Bracelet Color Material Info Number : Google
# Maurice Laxroix Strap|Bracelet Color Material Buckle Info Number : Site 
MANUFACTURER = "Jos Von Arx"
CATEGORY = "STRAPS>MAURICE LACROIX"


def cleanup(wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # print("Row to clean is: ", row_num)
    products_sheet.delete_rows(row_num)


def open_product_attributes(input_file):
    # The tsv will be converted to a list of dictionaries
    # Now we index as [prod_num]["type"]
    new_attrs = open(input_file, encoding = 'utf-8')
    new_attrs = [line.replace('\n','') for line in new_attrs] 
    new_attrs = [line.split('\t') for line in new_attrs]
    attr_names = new_attrs.pop(0)

    attrs_dicts = [{} for prod in new_attrs]
    for prod_ind in range(len(new_attrs)):
        for i in range(len(new_attrs[prod_ind])):
            attrs_dicts[prod_ind][attr_names[i]] = new_attrs[prod_ind][i]
    
    return attrs_dicts


def load_pickle_obj(file):
    import pickle

    with open(file, 'rb') as f:
        return pickle.load(f)

# XLSX modifier functions
def add_empty_product(product_info, wb):
    products_sheet = wb['Products']
    products_sheet.append(['' for i in range(products_sheet.max_column)])
    row_num = products_sheet.max_row

    # Write the new product code
    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1
    products_sheet['A' + str(row_num)] = curr_product_id
    
    print("Insert new product with ID {} in row {} with Model: {}".format( \
        curr_product_id, row_num, product_info["number"]))


def process_attr_data(product, attr_grp):
    for attr in COLUMN_TRANSF_RULES[attr_grp]:
        # First create the list(of lists) based on the priorities
        transformations_ordered = [[] for i in range(len(COLUMN_TRANSF_RULES[attr_grp][attr]) + 1000)] 
        
        for word, transf in COLUMN_TRANSF_RULES[attr_grp][attr].items():
            transformations_ordered[transf[-1]].append(word)
            # [ ['STEEL','S.S'], ['Ceramic'], ... ]
            
        # Clean empty lists
        transformations_ordered = [lst for lst in transformations_ordered if lst]
        # Watch out for no data
        try:
            if not product[attr]:
                product[attr] = ""
                continue
        except:
            product[attr] = ""


        new_val_el = ""
        new_val_en = ""
        for transf in transformations_ordered:
            possible_words = [word for word in transf]

            match_word = ""
            # Now check if any of the words are in the text
            for word in possible_words:
                if word == "__*":
                    match_word = word
                    break

                if "__color" in word:
                    colors_vals = get_color(word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    if colors_vals:
                        match_word = word
                        break

                elif word.upper() in product[attr].upper():
                    match_word = word
                    break

            # Match with the first word from the list
            if match_word:
                new_val_en += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][match_word][0]
                new_val_el += ' ' + COLUMN_TRANSF_RULES[attr_grp][attr][match_word][1]

                # If the pattern matched we have the colors that need to be replaced
                if "__color" in match_word:
                    colors_vals = get_color(match_word.lower(), replace_chars("/|", product[attr].lower(), " "))
                    new_val_el = new_val_el.replace("__color", colors_vals[1])
                    new_val_en = new_val_en.replace("__color", colors_vals[0])
        
        if not new_val_el:
            product[attr] = [product[attr], product[attr]]
        elif new_val_el == " __clear":
            # Special commands
            product[attr] = ["", ""]
        else:
            # Remove the first space
            new_val_el = new_val_el[1:][0].upper() + new_val_el[2:]
            new_val_en = new_val_en[1:][0].upper() + new_val_en[2:]
            if new_val_el[-1] == ',':
                new_val_el = new_val_el[:-1]
                new_val_en = new_val_en[:-1]
            product[attr] = [new_val_el, new_val_en]

    # Lastly all attributes that weren't filtered, convert them to [greek, engl] format
    for attr in product:
        if not isinstance(product[attr], list) and attr in COLUMN_TRANSF_RULES[attr_grp]:
            product[attr] = [product[attr], product[attr]]

    return product        


def static_pre_processing(product_info, attr_grp):
    # Apply rules to the data before inserting it
    product_info['name_material'] = product_info['ΥΛΙΚΟ']

    return product_info


def static_post_processing(product_info, attr_grp):  
    product_info['ΠΛΗΡΟΦΟΡΙΕΣ'] = ['', '']
    if product_info['info gr']:
        product_info['ΠΛΗΡΟΦΟΡΙΕΣ'] = [product_info['info gr'], product_info['info en']]

    product_info['ΥΛΙΚΟ'][0] = product_info['ΥΛΙΚΟ'][0].replace(',', ' και')
    product_info['ΥΛΙΚΟ'][1] = product_info['ΥΛΙΚΟ'][1].replace(',', ' and')

    product_info['brand'] = product_info['brand'].title()
    product_info['category_short'] = product_info['category'].split('>')[-1].lower()[:-1]

    return product_info


def add_attributes(product_info, wb):
    products_sheet = wb['Products']
    attributes_sheet = wb['ProductAttributes']
    row_num = products_sheet.max_row
    attr_row_num = attributes_sheet.max_row + 1

    (categories_to_attribute, __, attributes_dict) = load_pickle_obj('pkl_files/attributes.pkl')

    attr_grp = categories_to_attribute[product_info['category']]
    attributes = attributes_dict[attr_grp]

    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    # Data Processing
    attribute_info = static_pre_processing(product_info, attr_grp)
    attribute_info = process_attr_data(attribute_info, attr_grp)
    attribute_info = static_post_processing(attribute_info, attr_grp)

    i = 0
    for attr in attributes:
        # If attribute is empty, don't insert the row
        if attr not in attribute_info or attribute_info[attr][0] == "":
            continue
        attributes_sheet.append(['' for j in range(attributes_sheet.max_column)])
        
        attributes_sheet['C' + str(attr_row_num + i)] = attr
        attributes_sheet['A' + str(attr_row_num + i)] = curr_product_id
        attributes_sheet['B' + str(attr_row_num + i)] = attr_grp

        if attr in attribute_info:
            attributes_sheet['D' + str(attr_row_num + i)] = attribute_info[attr][0]
            attributes_sheet['E' + str(attr_row_num + i)] = attribute_info[attr][1]
        else:
            attributes_sheet['D' + str(attr_row_num + i)] = ''
            attributes_sheet['E' + str(attr_row_num + i)] = ''
        i += 1

    print(attribute_info['category_short'])
    print(attr_grp)


def add_product_name_and_meta_title(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Create product name
    if "bracelet" in product_info['category_short']:
        title_el = product_info['brand'] + ' Ανδρικό Βραχιόλι από ' \
                 + product_info['ΥΛΙΚΟ'][0].title() + ', ' + product_info['number']
        title_en = product_info['brand'] + ' Mens Bracelet made of ' \
                 + product_info['ΥΛΙΚΟ'][1].title() + ', ' + product_info['number']

    if "pen" in product_info['category_short']:
        title_el = product_info['brand'] + ' Στυλό από' \
                 + product_info['name_material'][0].title() + ', ' + product_info['number']
        title_en = product_info['brand'] + ' Pen made of ' \
                 + product_info['name_material'][1].title() + ', ' + product_info['number']

    if "wallet" in product_info['category_short']:
        title_el = product_info['brand'] + ' Ανδρικό Δερμάτινο Πορτοφόλι, ' + product_info['number']
        title_en = product_info['brand'] + ' Mens Leather Wallet, ' + product_info['number']

    if "card holder" in product_info['category_short']:
        title_el = product_info['brand'] + ' Δερμάτινη Θήκη για Κάρτες ' + product_info['number']
        title_en = product_info['brand'] + ' Leather Card Holder ' + product_info['number']

    meta_title_el = title_el + ' | Eurotimer'
    meta_title_el = title_el + ' | Eurotimer'
    meta_title_en = title_en + ' | Eurotimer'

    # Write product name
    products_sheet['B' + str(row_num)] = title_el
    products_sheet['C' + str(row_num)] = title_en

    # Write meta titles
    products_sheet['AG' + str(row_num)] = meta_title_el
    products_sheet['AH' + str(row_num)] = meta_title_en


def add_description(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    if "bracelet" in product_info['category_short']:
        descr_el = 'Ανδρικό βραχιόλι από ' + product_info['ΥΛΙΚΟ'][0].lower() \
                 + ' της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Mens bracelet made of ' + product_info['ΥΛΙΚΟ'][1].lower() \
                 + ' by ' + product_info['brand'] + '. '

    if "pen" in product_info['category_short']:
        descr_el = 'Στυλό από ' + product_info['name_material'][0].lower() \
                 + ' της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Pen made of ' + product_info['name_material'][1].lower() \
                 + ' by ' + product_info['brand'] + '. '


    if "wallet" in product_info['category_short']:
        descr_el = 'Ανδρικό πορτοφόλι από ' + product_info['ΥΛΙΚΟ'][0].lower() \
                 + ' της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Mens wallet made of ' + product_info['ΥΛΙΚΟ'][1].lower() \
                 + ' by ' + product_info['brand'] + '. '

    if "card holder" in product_info['category_short']:
        descr_el = 'Ανδρικό πορτοφόλι από ' + product_info['ΥΛΙΚΟ'][0].lower() \
                 + ' της εταιρίας ' + product_info['brand'] + '. '
        descr_en = 'Mens wallet made of ' + product_info['ΥΛΙΚΟ'][1].lower() \
                 + ' by ' + product_info['brand'] + '. '

    if product_info['info gr']:
        descr_el += product_info['info gr']
        descr_en += product_info['info en']

    meta_descr_el = descr_el
    meta_descr_en = descr_en

    # Write description
    products_sheet['AE' + str(row_num)] = '<p>' + descr_el + '</p>'
    products_sheet['AF' + str(row_num)] = '<p>' + descr_en + '</p>'

    # Write meta description
    products_sheet['AI' + str(row_num)] = meta_descr_el
    products_sheet['AJ' + str(row_num)] = meta_descr_en


def add_SEO(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    SEO = (product_info['brand'] + '-' + product_info['category_short'] + '-' \
        + product_info['number']).replace(' ', '_')

    # Write SEO
    products_sheet['AD' + str(row_num)] = SEO


def add_model(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get model
    model = product_info["number"]

    # Write model
    products_sheet['M' + str(row_num)] = model


def add_price(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Get price
    price = product_info["price"]

    # Write price
    products_sheet['Q' + str(row_num)] = price


def add_manufacturer(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    # Check manufacturer
    manuf = product_info['brand']
    if manuf == '':
        print('Warning: invalid manufacturer name!')

    # Write manufacturer
    products_sheet['N' + str(row_num)] = manuf


def add_category(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    categories_dict = load_pickle_obj('pkl_files/categories.pkl')

    # Find category number from dictionary
    categ = closest_match(product_info['category'], categories_dict)

    # Also add the parent categories
    broken_category = categ.split(">")
    parent_categ = broken_category[0]
    categ_val = str(categories_dict[parent_categ])

    for i in range(1, len(broken_category)):
        parent_categ = parent_categ + '>' + broken_category[i]
        categ_val += "," + str( categories_dict[parent_categ] )


    # Write category number
    products_sheet['D' + str(row_num)] = categ_val


def add_status(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row

    if not product_info['hidden']:
        products_sheet['AB' + str(row_num)] = 'true'
    else:
        products_sheet['AB' + str(row_num)] = 'false'


def add_image(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
              + '/' + product_info["number"] + "a.jpg"

    # Write image directory
    products_sheet['O' + str(row_num)] = image_dir

    # Extra image
    sheet = wb["AdditionalImages"]
    sheet.append(['' for i in range(sheet.max_column)])
    row = sheet.max_row 

    last_product_id = products_sheet['A' + str(row_num - 1)].value
    curr_product_id = last_product_id + 1

    image_dir = 'catalog/product/' + product_info['category'].replace('>', '/') \
              + '/' + product_info["number"] + "b.jpg"
    sheet['A' + str(row)] = curr_product_id
    sheet['B' + str(row)] = image_dir
    sheet['C' + str(row)] = 1


def add_misc(product_info, wb):
    products_sheet = wb['Products']
    row_num = products_sheet.max_row
    
    products_sheet['L'  + str(row_num)] = product_info["stock"]      #quantity
    products_sheet['P'  + str(row_num)] = 'yes'  #shipping
    products_sheet['R'  + str(row_num)] = 0      #points
    products_sheet['S'  + str(row_num)] = '2018-09-25 18:00:00' #date_added
    products_sheet['T'  + str(row_num)] = '2018-09-25 18:00:00'#date_modified
    products_sheet['U'  + str(row_num)] = '2018-09-25' #date_available
    products_sheet['V'  + str(row_num)] = 0      #weight
    products_sheet['W'  + str(row_num)] = 'kg'   #weight_unit
    products_sheet['X'  + str(row_num)] = 0      #length
    products_sheet['Y'  + str(row_num)] = 0      #width
    products_sheet['Z'  + str(row_num)] = 0      #height
    products_sheet['AA' + str(row_num)] = 'cm'   #length_unit
    products_sheet['AC' + str(row_num)] = 0      #tax_class_id
    products_sheet['AM' + str(row_num)] = 6      #stock_status_id
    products_sheet['AN' + str(row_num)] = 0      #store_ids
    products_sheet['AS' + str(row_num)] = 1      #sort_order
    products_sheet['AT' + str(row_num)] = 'true' #subtract
    products_sheet['AU' + str(row_num)] = 1      #minimum


if __name__ == '__main__':

    if len(argv) < 3:
        print('arg1: specs.tsv | arg2: products.xlsx')
        exit(1)

    SPECS_TSV = argv[1]
    products_xlsx = argv[2]

    wb = load_workbook(products_xlsx)
    new_products = open_product_attributes(SPECS_TSV)
    products_count = 0
    

    # Iterate the inputs
    for product in new_products:
        if not product["category"]: continue
        add_empty_product(product, wb)
        add_attributes(product, wb)
        add_product_name_and_meta_title(product, wb)
        add_description(product, wb)
        add_SEO(product, wb)
        add_model(product, wb)
        add_price(product, wb)
        add_manufacturer(product, wb)
        add_category(product, wb)
        add_status(product, wb)
        add_image(product, wb)
        add_misc(product, wb)
        products_count += 1
        wb.save(products_xlsx)

    # Cleanup and Save to file
    print("Added {} products.".format(products_count))
    cleanup(wb)
    wb.save(products_xlsx)
